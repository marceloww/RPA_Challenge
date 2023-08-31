from botcity.web import WebBot, Browser
import pandas as pd
import time
# Uncomment the line below for integrations with BotMaestro
# Using the Maestro SDK
from botcity.maestro import *
from botcity.web import By
import openpyxl



class Bot(WebBot):
    def action(self, execution=None):
        self.headless = False
        self.driver_path = "./drivers/chromedriver.exe"
        self.browse("https://rpachallenge.com/?lang=EN")
        self.wait(3000)
        # Caminho das planilhas
        tabela = pd.read_excel("C:/Users/marce/OneDrive/Área de Trabalho/RPA Challenge/Desafio RPA/Plan 1.xlsx")
        num_linhas = len(tabela)
        linha_atual = 0

        # Lista com os nomes das planilhas
        nomes_planilhas = ["Plan 1.xlsx", "Plan 2.xlsx", "Plan 3.xlsx"]

        for nome_planilha in nomes_planilhas:
            # Caminho da planilha
            planilha_path = f"C:/Users/marce/OneDrive/Área de Trabalho/RPA Challenge/Desafio RPA/{nome_planilha}"

            # Carregar planilha com pandas
            tabela = pd.read_excel(planilha_path)
            num_linhas = len(tabela)
            linha_atual = 0

            for i in range(linha_atual, num_linhas):
                try:
                    lancado_status = tabela.loc[i, 'Lançado']
                    if lancado_status == 'OK':
                        print(f"Essa linha {i + 1} da {nome_planilha} ja foi lançada!")
                        continue
                    first_name = tabela.loc[i, 'First Name']
                    last_name = tabela.loc[i, 'Last Name']
                    company_name = tabela.loc[i, 'Company Name']
                    role = tabela.loc[i, 'Role in Company']
                    address = tabela.loc[i, 'Address']
                    email = tabela.loc[i, 'Email']
                    phone = tabela.loc[i, 'Phone Number']

                    click_name = self.find_element(selector="//label[contains(text(), 'First Name')]/following-sibling::input", by=By.XPATH)
                    if click_name:
                        click_name.click()
                        self.paste(str(first_name))


                    click_name = self.find_element(selector="//label[contains(text(), 'Last Name')]/following-sibling::input", by=By.XPATH)
                    if click_name:
                        click_name.click()
                        self.paste(str(last_name))


                    click_name = self.find_element(selector="//label[contains(text(), 'Company Name')]/following-sibling::input", by=By.XPATH)
                    if click_name:
                        click_name.click()
                        self.paste(str(company_name))


                    click_name = self.find_element(selector="//label[contains(text(), 'Role in Company')]/following-sibling::input", by=By.XPATH)
                    if click_name:
                        click_name.click()
                        self.paste(str(role))


                    click_name = self.find_element(selector="//label[contains(text(), 'Address')]/following-sibling::input", by=By.XPATH)
                    if click_name:
                        click_name.click()
                        self.paste(str(address))


                    click_name = self.find_element(selector="//label[contains(text(), 'Email')]/following-sibling::input", by=By.XPATH)
                    if click_name:
                        click_name.click()
                        self.paste(str(email))


                    click_name = self.find_element(selector="//label[contains(text(), 'Phone Number')]/following-sibling::input", by=By.XPATH)
                    if click_name:
                        click_name.click()
                        self.paste(str(phone))


                    click_submit = self.find_element(selector='/html/body/app-root/div[2]/app-rpa1/div/div[2]/form/input', by=By.XPATH)
                    if click_submit:
                        click_submit.click()



                    # Abre o arquivo Excel com openpyxl para editar sem perder a formatação
                    workbook = openpyxl.load_workbook(planilha_path)
                    sheet = workbook.active

                    # Localiza a coluna 'Lançado' pelo nome
                    coluna_lancado = None
                    for col in sheet.iter_cols(min_col=1, max_col=sheet.max_column, min_row=1, max_row=1):
                        for cell in col:
                            if cell.value == 'Lançado':
                                coluna_lancado = cell.column
                                break
                        if coluna_lancado:
                            break

                    if coluna_lancado:
                        # Preenche a célula na linha atual
                        coluna_atual = sheet.cell(row=i + 2, column=coluna_lancado)
                        coluna_atual.value = 'OK'

                    # Salva as alterações no arquivo
                    workbook.save(planilha_path)

                except Exception as e:
                    print(f"Ocorreu um erro: {e}")

            else:
                print(f"Todos os registros da planilha {nome_planilha} processados com sucesso!")

        self.stop_browser()

if __name__ == '__main__':
    Bot.main()
